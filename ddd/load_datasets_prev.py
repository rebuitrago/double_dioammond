"""
Load the Track-B research files into `observation`:
  GII, PCI, FSI, UNCTAD FDI  (original)
  + V-Dem polyarchy           (added: WIDE_CONFIG entry — same shape as the others)
  + Euromonitor inflation & unemployment (added: load_euromonitor() — Excel + forecast flag)

WBES is NOT here — pulled automatically by ingest.py (WDI IC.FRM.* codes).
QoG was dropped (redundant with WGI). Euromonitor Real GDP Growth is NOT here —
it is a relay onto criterion #6 and is loaded by the (later) relay step.

USAGE
    pip install pycountry openpyxl       # one-time
    export SUPABASE_URL=...  SUPABASE_SECRET_KEY=...  SUPABASE_SERVICE_KEY=...
    python -m ddd.load_datasets

Notes
  * GII uses ISO3 already; PCI/FSI/V-Dem use country names (auto-resolved);
    UNCTAD uses UN M49 numeric codes (auto-resolved). Rows for economies not in
    your `country` table are skipped.
  * Euromonitor: the 'Charting - Annual' sheet, Baseline scenario, with the
    Historic/Forecast band setting value_type, and leading 0s treated as missing.
  * Header strings below must match your files EXACTLY (caps, spaces, colons).
"""

from __future__ import annotations
import os
import pandas as pd
from ddd.connectors import upsert_observations

DRY_RUN = False  # set by --dry-run: reads the DB but writes nothing

def _upsert(rows):
    """Write rows unless DRY_RUN is set (then just report the count)."""
    if DRY_RUN:
        return len(rows)
    return upsert_observations(rows)

def _dedupe(rows):
    """Keep one row per (indicator_id, country_iso3, year) — last wins."""
    seen = {}
    for r in rows:
        seen[(r['indicator_id'], r['country_iso3'], r['year'])] = r
    return list(seen.values())

# ---- EDIT THESE PATHS -----------------------------------------------------
GII_FILE    = "data/gii.csv"
PCI_FILE    = "data/pci.csv"
FSI_FILE    = "data/fsi.csv"
UNCTAD_FILE = "data/unctad_fdi.csv"     # set to None to skip UNCTAD
VDEM_FILE   = "data/v-dem_poliarchy.csv"
# Euromonitor: (xlsx path, source_code). Real GDP Growth is intentionally omitted
# (it's a relay onto criterion #6, loaded later, not a standalone criterion).
EUROMONITOR_FILES = [
    ("data/Euromonitor_Inflation.XLSX",    "EUROMONITOR.INFLATION"),
    ("data/Euromonitor_Unemployment.XLSX", "EUROMONITOR.UNEMPLOYMENT"),
]
EUROMONITOR_VINTAGE  = "Euromonitor 2026-05"   # match the file's 'Last updated'
EUROMONITOR_MAX_YEAR = 2031                     # align horizon with WEO/EIU
# EIU Business Environment: one .xlsx, 3 sheets (one per sub-component). The
# actual/estimate/forecast flag is encoded in CELL FORMATTING, not a column:
#   forecast = light-blue fill (FFD9E8F2);  estimate = blue font (FF00588D);
#   actual   = black font. (Row-banding greys are cosmetic, not flags.)
EIU_FILE    = "data/EIU_dd.xlsx"
EIU_VINTAGE = "EIU 2026-04"     # matches the file's 'Published' date (08-04-2026)
EIU_SHEET_TO_CODE = {
    "financing":              "EIU.BE.FINANCING",
    "labour":                 "EIU.BE.LABOUR_SKILLS",
    "policy towards private": "EIU.BE.PRIV_ENTERPRISE",
}
EIU_FORECAST_FILL = "FFD9E8F2"
EIU_ESTIMATE_FONT = "FF00588D"

# --- Relays: gap-fill an existing criterion with a fresher source for the years
#     the primary (WDI) lacks. No overlap with the primary, so relayed cells slot
#     into the one-row-per-cell table; each is tagged with RELAY_VINTAGE_SUFFIX in
#     source_vintage and EXCLUDED when computing the primary frontier (so re-runs
#     are idempotent). The relay fills only years strictly after the frontier.
RELAY_VINTAGE_SUFFIX = " (relay)"
EUROMONITOR_GROWTH_FILE = "data/Euromonitor Real GDP Growth.XLSX"   # note: spaces in the filename
RELAYS = [
    {
        "target_code": "NY.GDP.MKTP.KD.ZG",   # WDI 'GDP growth (annual %)' = criterion #6 (Demand size)
        "label": "Euromonitor real GDP growth -> WDI GDP growth (#6)",
        "file": EUROMONITOR_GROWTH_FILE,
        "vintage": "Euromonitor 2026-05",
    },
]
# ---------------------------------------------------------------------------

# source_code -> (file, country_col, year_col, value_col, country_kind)
WIDE_CONFIG: dict[str, tuple] = {
    # WGI governance (built from the World Bank Excel into data/wgi.csv; ISO3 codes)
    "GE.EST": ("data/wgi.csv", "iso3", "year", "GE.EST", "iso3"),
    "RQ.EST": ("data/wgi.csv", "iso3", "year", "RQ.EST", "iso3"),
    "RL.EST": ("data/wgi.csv", "iso3", "year", "RL.EST", "iso3"),
    "CC.EST": ("data/wgi.csv", "iso3", "year", "CC.EST", "iso3"),
    # GII (scores 0-100)
    "GII.HUMAN_CAPITAL_RESEARCH":  (GII_FILE, "country", "year", "Human capital and research", "name"),
    "GII.INFRASTRUCTURE":          (GII_FILE, "country", "year", "Infrastructure", "name"),
    "GII.MARKET_SOPHISTICATION":   (GII_FILE, "country", "year", "Market sophistication", "name"),
    "GII.BUSINESS_SOPHISTICATION": (GII_FILE, "country", "year", "Business sophistication", "name"),
    "GII.KNOWLEDGE_TECH_OUTPUTS":  (GII_FILE, "country", "year", "Knowledge and technology outputs", "name"),
    "GII.KNOWLEDGE_DIFFUSION":     (GII_FILE, "country", "year", "Knowledge diffusion", "name"),
    # PCI (names; note year_ trailing underscore, pci_Institutions capital I)
    "PCI.HUMAN_CAPITAL":   (PCI_FILE, "economy", "year_", "pci_hum_cap", "name"),
    "PCI.NATURAL_CAPITAL": (PCI_FILE, "economy", "year_", "pci_nat_cap", "name"),
    "PCI.ENERGY":          (PCI_FILE, "economy", "year_", "pci_energy", "name"),
    "PCI.TRANSPORT":       (PCI_FILE, "economy", "year_", "pci_transp", "name"),
    "PCI.ICT":             (PCI_FILE, "economy", "year_", "pci_ict", "name"),
    "PCI.INSTITUTIONS":    (PCI_FILE, "economy", "year_", "pci_Institutions", "name"),
    "PCI.PRIVATE_SECTOR":  (PCI_FILE, "economy", "year_", "pci_priv_sect", "name"),
    # FSI (names; scale 0-10)
    "FSI.C1_SECURITY":     (FSI_FILE, "Country", "Year", "C1: Security Apparatus", "name"),
    "FSI.C2_ELITES":       (FSI_FILE, "Country", "Year", "C2: Factionalized Elites", "name"),
    "FSI.P1_LEGITIMACY":   (FSI_FILE, "Country", "Year", "P1: State Legitimacy", "name"),
    "FSI.P2_PUBLIC_SVC":   (FSI_FILE, "Country", "Year", "P2: Public Services", "name"),
    "FSI.P3_HUMAN_RIGHTS": (FSI_FILE, "Country", "Year", "P3: Human Rights", "name"),
    "FSI.E3_BRAIN_DRAIN":  (FSI_FILE, "Country", "Year", "E3: Human Flight and Brain Drain", "name"),
    # V-Dem polyarchy (names; 0-1 index; all values are ACTUAL incl. 2025)
    "VDEM.V2X_POLYARCHY":  (VDEM_FILE, "country_name", "year", "v2x_polyarchy", "name"),
}

# Country-name spellings that pycountry's fuzzy match gets wrong or ambiguous.
OVERRIDES = {
    "Korea, Rep.": "KOR", "Korea, Republic of": "KOR", "Republic of Korea": "KOR",
    "South Korea": "KOR", "Korea, Dem. Rep.": "PRK", "North Korea": "PRK",
    "Russia": "RUS", "Russian Federation": "RUS",
    "Turkey": "TUR", "Türkiye": "TUR", "Turkiye": "TUR",
    "T√ºrkiye": "TUR",                       # ADDED: V-Dem export mojibake of 'Türkiye'
    "Kosovo": "XKX",                         # ADDED: pin so it can't fuzzy-merge into Serbia (SRB)
    "Vietnam": "VNM", "Viet Nam": "VNM",
    "Iran": "IRN", "Iran, Islamic Rep.": "IRN", "Iran (Islamic Republic of)": "IRN",
    "Bolivia": "BOL", "Bolivia (Plurinational State of)": "BOL",
    "Venezuela": "VEN", "Venezuela, RB": "VEN", "Venezuela (Bolivarian Republic of)": "VEN",
    "Egypt": "EGY", "Egypt, Arab Rep.": "EGY",
    "Syria": "SYR", "Syrian Arab Republic": "SYR",
    "Laos": "LAO", "Lao PDR": "LAO", "Lao People's Democratic Republic": "LAO",
    "Tanzania": "TZA", "Tanzania, United Rep.": "TZA",
    "Moldova": "MDA", "Republic of Moldova": "MDA",
    "Czechia": "CZE", "Czech Republic": "CZE",
    "Slovakia": "SVK", "Slovak Republic": "SVK",
    "Brunei": "BRN", "Brunei Darussalam": "BRN",
    "Cote d'Ivoire": "CIV", "Côte d'Ivoire": "CIV", "Ivory Coast": "CIV",
    "Congo, Dem. Rep.": "COD", "Democratic Republic of the Congo": "COD",
    "Congo, Rep.": "COG", "Republic of the Congo": "COG", "Congo": "COG",
    "Kyrgyz Republic": "KGZ", "Kyrgyzstan": "KGZ",
    "Gambia, The": "GMB", "Gambia": "GMB", "Bahamas, The": "BHS",
    "Hong Kong": "HKG", "Hong Kong SAR, China": "HKG",
    "Macao": "MAC", "Macao SAR, China": "MAC",
    "United States": "USA", "United States of America": "USA",
    "Cape Verde": "CPV", "Cabo Verde": "CPV",
    "Eswatini": "SWZ", "Swaziland": "SWZ",
    "North Macedonia": "MKD", "Macedonia, FYR": "MKD",
    "Slovak Republic": "SVK", "Yemen, Rep.": "YEM",
}

_FUZZY_CACHE: dict[str, str | None] = {}


def name_to_iso3(name) -> str | None:
    n = str(name).strip()
    if n in OVERRIDES:
        return OVERRIDES[n]
    if n in _FUZZY_CACHE:
        return _FUZZY_CACHE[n]
    iso3 = None
    try:
        import pycountry
        c = (pycountry.countries.get(name=n)
             or pycountry.countries.get(official_name=n)
             or pycountry.countries.get(common_name=n))
        if not c:
            hits = pycountry.countries.search_fuzzy(n)
            c = hits[0] if hits else None
        iso3 = c.alpha_3 if c else None
    except Exception:
        iso3 = None
    _FUZZY_CACHE[n] = iso3
    return iso3


def m49_to_iso3(code) -> str | None:
    try:
        import pycountry
        c = pycountry.countries.get(numeric=str(int(code)).zfill(3))
        return c.alpha_3 if c else None
    except Exception:
        return None


def _client():
    from supabase import create_client
    return create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_SECRET_KEY"])


def _read_all(client, table, select):
    rows, start = [], 0
    while True:
        resp = client.table(table).select(select).range(start, start + 999).execute()
        b = resp.data or []
        rows.extend(b)
        if len(b) < 1000:
            break
        start += 1000
    return rows


def _code_to_id(client) -> dict[str, int]:
    return {r["source_code"]: r["id"]
            for r in _read_all(client, "indicator", "id,source_code")
            if r.get("source_code")}


def _country_iso3(client) -> set[str]:
    return {r["iso3"] for r in _read_all(client, "country", "iso3")}


def _gdp_lookup(client, gdp_id) -> dict[tuple, float]:
    if gdp_id is None:
        return {}
    out = {}
    start = 0
    while True:
        resp = (client.table("observation")
                .select("country_iso3,year,raw_value")
                .eq("indicator_id", gdp_id).range(start, start + 999).execute())
        b = resp.data or []
        for r in b:
            out[(r["country_iso3"], int(r["year"]))] = float(r["raw_value"])
        if len(b) < 1000:
            break
        start += 1000
    return out


def load_wide(client, code_to_id, valid) -> int:
    total = 0
    for code, (path, ccol, ycol, vcol, kind) in WIDE_CONFIG.items():
        ind = code_to_id.get(code)
        if ind is None:
            print(f"  ! {code}: not in DB (run migrations first). skipped")
            continue
        if not os.path.exists(path):
            print(f"  ! {code}: file {path} not found. skipped")
            continue
        df = pd.read_csv(path, encoding="utf-8-sig")   # utf-8-sig: tolerates a BOM (V-Dem export has one)
        if vcol not in df.columns:
            print(f"  ! {code}: column '{vcol}' not in {os.path.basename(path)}. "
                  f"Check exact header. skipped")
            continue
        rows, skipped = [], 0
        for rec in df.to_dict("records"):
            raw = rec.get(vcol)
            if isinstance(raw, str) and raw.strip() in (".", "", "..", "NA", "n/a"): continue
            if raw is None or pd.isna(raw):
                continue
            iso3 = rec[ccol] if kind == "iso3" else name_to_iso3(rec[ccol])
            if not iso3 or iso3 not in valid:
                skipped += 1
                continue
            try:
                yr = int(rec[ycol])
            except (ValueError, TypeError):
                continue
            rows.append({"indicator_id": ind, "country_iso3": str(iso3),
                         "year": yr, "raw_value": float(raw),
                         "source_vintage": code.split(".")[0] + " import"})
        rows = _dedupe(rows)
        n = _upsert(rows)
        total += n
        note = f" ({skipped} rows skipped: not in country set)" if skipped else ""
        print(f"  + {code}: {n}{note}")
    return total


def load_euromonitor(client, code_to_id, valid) -> int:
    """Load Euromonitor annual series (Inflation, Unemployment) with a value_type
    flag from the Historic/Forecast band. Custom because the file is a multi-sheet
    Excel, wide across years, with the forecast boundary encoded in a header row."""
    import openpyxl
    SHEET = "Charting - Annual"
    HDR, DATA0 = 5, 6
    C_COUNTRY, C_IND, C_SCEN = 2, 3, 4
    total = 0
    for path, code in EUROMONITOR_FILES:
        ind = code_to_id.get(code)
        if ind is None:
            print(f"  ! {code}: not in DB (run migration_03 first). skipped")
            continue
        if not os.path.exists(path):
            print(f"  ! {code}: file {path} not found. skipped")
            continue
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if SHEET not in wb.sheetnames:
            print(f"  ! {code}: no '{SHEET}' sheet in {os.path.basename(path)}. skipped")
            continue
        ws = wb[SHEET]
        # year columns (row 5) + forecast boundary (row 4 'Forecast' label)
        year_col, fc_year = {}, None
        for c in range(C_SCEN + 1, ws.max_column + 1):
            hv = ws.cell(row=HDR, column=c).value
            if isinstance(hv, (int, float)):
                year_col[c] = int(hv)
            band = ws.cell(row=4, column=c).value
            if band and "Forecast" in str(band):
                fc_year = ws.cell(row=HDR, column=c).value
        if not year_col or fc_year is None:
            print(f"  ! {code}: could not find year headers / Forecast band. skipped")
            continue
        fc_year = int(fc_year)

        rows, skipped, dropped0 = [], 0, 0
        covered = set()
        for r in range(DATA0, ws.max_row + 1):
            cn = ws.cell(row=r, column=C_COUNTRY).value
            scen = ws.cell(row=r, column=C_SCEN).value
            if not cn or not scen or "Baseline" not in str(scen):
                continue                      # skips footer + non-baseline rows
            iso3 = name_to_iso3(cn)
            if not iso3 or iso3 not in valid:
                skipped += 1
                continue
            for c, yr in year_col.items():
                if yr > EUROMONITOR_MAX_YEAR:
                    continue
                v = ws.cell(row=r, column=c).value
                if v is None or not isinstance(v, (int, float)):
                    continue
                if v == 0:                    # leading 0 == missing sentinel (verified)
                    dropped0 += 1
                    continue
                covered.add(iso3)
                rows.append({"indicator_id": ind, "country_iso3": iso3, "year": yr,
                             "raw_value": float(v), "source_vintage": EUROMONITOR_VINTAGE,
                             "value_type": "forecast" if yr >= fc_year else "actual"})
        rows = _dedupe(rows)
        n_fc = sum(1 for x in rows if x["value_type"] == "forecast")
        n = _upsert(rows)
        total += n
        missing = sorted(valid - covered)
        print(f"  + {code}: {n} ({n - n_fc} actual / {n_fc} forecast, fc>= {fc_year}); "
              f"{len(covered)}/{len(valid)} countries; {dropped0} zeros dropped")
        if missing:
            print(f"      not covered by Euromonitor ({len(missing)}): {missing}")
    return total


def _eiu_year(v):
    try:
        y = int(float(str(v).strip()))
        return y if 1900 <= y <= 2100 else None
    except (ValueError, TypeError):
        return None


def _eiu_value_type(cell):
    """actual / estimate / forecast from the cell's formatting (see EIU legend)."""
    fill = cell.fill.fgColor.rgb if (cell.fill and cell.fill.patternType) else None
    if isinstance(fill, str) and fill.upper() == EIU_FORECAST_FILL:
        return "forecast"
    fc = cell.font.color.rgb if (cell.font and cell.font.color) else None
    if isinstance(fc, str) and fc.upper() == EIU_ESTIMATE_FONT:
        return "estimate"
    return "actual"


def load_eiu(client, code_to_id, valid) -> int:
    """Load the 3 EIU Business Environment sub-components. Reads cell colours to
    set value_type, so it must NOT use read_only mode (which drops styles)."""
    import openpyxl
    if not os.path.exists(EIU_FILE):
        print(f"  ! EIU: file {EIU_FILE} not found. skipped")
        return 0
    wb = openpyxl.load_workbook(EIU_FILE)   # default load: keeps fill/font colours
    total = 0
    for sh in wb.sheetnames:
        code = next((v for k, v in EIU_SHEET_TO_CODE.items() if k in sh.lower()), None)
        if code is None:
            print(f"  ! EIU sheet {sh!r}: no source_code mapping. skipped")
            continue
        ind = code_to_id.get(code)
        if ind is None:
            print(f"  ! {code}: not in DB (run migration_03 first). skipped")
            continue
        ws = wb[sh]
        hdr = next((r for r in range(1, 15)
                    if str(ws.cell(row=r, column=1).value).strip() == "Geography"), None)
        if hdr is None:
            print(f"  ! {code}: no 'Geography' header in {sh!r}. skipped")
            continue
        year_col = {c: _eiu_year(ws.cell(row=hdr, column=c).value)
                    for c in range(9, ws.max_column + 1)
                    if _eiu_year(ws.cell(row=hdr, column=c).value)}
        rows, skipped, covered = [], 0, set()
        vt = {"actual": 0, "estimate": 0, "forecast": 0}
        for r in range(hdr + 1, ws.max_row + 1):
            g = ws.cell(row=r, column=1).value
            if not g:
                continue
            if any(k in str(g).lower() for k in
                   ("copyright", "economist", "exported", "updated", "\u00a9", "source:")):
                continue                       # footer guard
            iso3 = name_to_iso3(g)
            if not iso3 or iso3 not in valid:
                skipped += 1
                continue
            for c, yr in year_col.items():
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell.value, (int, float)):
                    continue
                kind = _eiu_value_type(cell)
                vt[kind] += 1
                covered.add(iso3)
                rows.append({"indicator_id": ind, "country_iso3": iso3, "year": yr,
                             "raw_value": float(cell.value),
                             "source_vintage": EIU_VINTAGE, "value_type": kind})
        rows = _dedupe(rows)
        n = _upsert(rows)
        total += n
        miss = sorted(valid - covered)
        print(f"  + {code}: {n} (actual {vt['actual']} / estimate {vt['estimate']} "
              f"/ forecast {vt['forecast']}); {len(covered)}/{len(valid)} countries")
        if miss:
            print(f"      not covered by EIU ({len(miss)}): {miss}")
    return total


def _euromonitor_annual_rows(path, valid):
    """(iso3, year, value, value_type) from a Euromonitor 'Charting - Annual' sheet.
    Same parsing as load_euromonitor; used by the growth relay. Exact-0 == missing."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Charting - Annual"]
    year_col, fc = {}, None
    for c in range(5, ws.max_column + 1):
        hv = ws.cell(row=5, column=c).value
        if isinstance(hv, (int, float)):
            year_col[c] = int(hv)
        b = ws.cell(row=4, column=c).value
        if b and "Forecast" in str(b):
            fc = int(ws.cell(row=5, column=c).value)
    out = []
    for r in range(6, ws.max_row + 1):
        cn = ws.cell(row=r, column=2).value
        sc = ws.cell(row=r, column=4).value
        if not cn or not sc or "Baseline" not in str(sc):
            continue
        iso = name_to_iso3(cn)
        if not iso or iso not in valid:
            continue
        for c, yr in year_col.items():
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, (int, float)) or v == 0:
                continue
            out.append((iso, yr, float(v), "forecast" if (fc and yr >= fc) else "actual"))
    return out, fc


def _primary_frontier(client, indicator_id):
    """Latest year for this indicator among NON-relay rows (the primary's frontier)."""
    start, maxyr = 0, None
    while True:
        resp = (client.table("observation").select("year,source_vintage")
                .eq("indicator_id", indicator_id).range(start, start + 999).execute())
        b = resp.data or []
        for r in b:
            if str(r.get("source_vintage") or "").endswith(RELAY_VINTAGE_SUFFIX):
                continue                       # ignore our own prior relay rows
            y = int(r["year"])
            if maxyr is None or y > maxyr:
                maxyr = y
        if len(b) < 1000:
            break
        start += 1000
    return maxyr


def load_relays(client, code_to_id, valid) -> int:
    """Gap-fill relays: extend a criterion past the primary's last year only."""
    total = 0
    for relay in RELAYS:
        code = relay["target_code"]
        ind = code_to_id.get(code)
        if ind is None:
            print(f"  ! relay [{relay['label']}]: target {code!r} not in catalog. skipped")
            continue
        frontier = _primary_frontier(client, ind)
        if frontier is None:
            print(f"  ! relay [{relay['label']}]: target #{ind} has no primary data "
                  f"(run ingest first). skipped")
            continue
        if not os.path.exists(relay["file"]):
            print(f"  ! relay [{relay['label']}]: file {relay['file']} not found. skipped")
            continue
        melt, fc = _euromonitor_annual_rows(relay["file"], valid)
        rows, covered = [], set()
        for iso, yr, val, vt in melt:
            if yr <= frontier or yr > EUROMONITOR_MAX_YEAR:
                continue                       # primary owns <= frontier; cap the horizon
            covered.add(iso)
            rows.append({"indicator_id": ind, "country_iso3": iso, "year": yr,
                         "raw_value": val,
                         "source_vintage": relay["vintage"] + RELAY_VINTAGE_SUFFIX,
                         "value_type": vt})
        rows = _dedupe(rows)
        n_fc = sum(1 for r in rows if r["value_type"] == "forecast")
        n = _upsert(rows)
        total += n
        print(f"  + RELAY {code} (#{ind}): {n} rows  [frontier {frontier} -> fills "
              f"{frontier + 1}+]  {n - n_fc} actual / {n_fc} forecast  "
              f"{len(covered)}/{len(valid)} countries")
    return total


def load_unctad(client, code_to_id, valid) -> int:
    in_id, out_id = code_to_id.get("UNCTAD.FDI.STOCK.IN"), code_to_id.get("UNCTAD.FDI.STOCK.OUT")
    if not in_id or not out_id:
        print("  ! UNCTAD indicators not in DB (run migration_02). skipped")
        return 0
    if not UNCTAD_FILE or not os.path.exists(UNCTAD_FILE):
        print(f"  ! UNCTAD file not found. skipped")
        return 0
    gdp = _gdp_lookup(client, code_to_id.get("NY.GDP.MKTP.CD"))
    if not gdp:
        print("  ! no GDP (NY.GDP.MKTP.CD) in observations — run ingest first. skipped")
        return 0
    df = pd.read_csv(UNCTAD_FILE)
    vcol = next((c for c in df.columns if "current prices" in c.lower()), None)
    if vcol is None:
        print("  ! UNCTAD: value column ('...current prices...') not found. skipped")
        return 0
    rows, dropped = [], 0
    for rec in df.to_dict("records"):
        flow = str(rec.get("Flow Label", rec.get("Flow", ""))).lower()
        if "stock" not in flow and str(rec.get("Flow", "")) != "9":
            continue  # keep stock only, drop flows
        iso3 = m49_to_iso3(rec.get("Economy"))
        if not iso3 or iso3 not in valid:   # drops World/regions + non-panel countries
            dropped += 1
            continue
        val = rec.get(vcol)
        if isinstance(val, str) and val.strip() in (".", "", "..", "NA", "n/a"): continue
        if val is None or pd.isna(val):
            continue
        try:
            yr = int(rec["Year"])
        except (ValueError, TypeError):
            continue
        g = gdp.get((iso3, yr))
        if not g:
            continue
        pct = float(val) * 1e6 / g * 100.0     # US$ millions -> % of GDP
        direction = str(rec.get("Direction Label", rec.get("Direction", ""))).lower()
        ind = in_id if ("in" in direction or str(rec.get("Direction")) == "1") else out_id
        rows.append({"indicator_id": ind, "country_iso3": iso3, "year": yr,
                     "raw_value": pct, "source_vintage": "UNCTAD FDI stock /GDP"})
    # dedupe + upsert ONCE, after the loop (previously these ran inside the loop,
    # re-uploading the whole growing list on every row -> very slow / appears frozen)
    rows = _dedupe(rows)
    n = _upsert(rows)
    print(f"  + UNCTAD inward+outward stock (% GDP): {n} ({dropped} aggregate/non-panel rows dropped)")
    return n


def main() -> None:
    import argparse
    global DRY_RUN
    ap = argparse.ArgumentParser(description="Load Track-B datasets into observation")
    ap.add_argument("--dry-run", action="store_true",
                    help="read the DB and parse files, but write nothing")
    DRY_RUN = ap.parse_args().dry_run
    if DRY_RUN:
        print("*** DRY RUN — reading and parsing only, no writes ***")
    client = _client()
    code_to_id = _code_to_id(client)
    valid = _country_iso3(client)
    print(f"Loading Track-B files into Supabase ({len(valid)} countries in panel)...")
    total = load_wide(client, code_to_id, valid)
    total += load_euromonitor(client, code_to_id, valid)
    total += load_eiu(client, code_to_id, valid)
    total += load_relays(client, code_to_id, valid)
    total += load_unctad(client, code_to_id, valid)
    verb = "Would upsert" if DRY_RUN else "Upserted"
    print(f"Done. {verb} {total} observations." +
          ("" if DRY_RUN else " Now re-run scoring with --scoring-method percentile."))


if __name__ == "__main__":
    main()
